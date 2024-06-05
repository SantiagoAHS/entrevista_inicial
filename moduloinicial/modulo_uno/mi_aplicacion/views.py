# mi_aplicacion/views.py
from django.shortcuts import render

def home(request):
    return render(request, 'home.html')

from django.shortcuts import render
from .forms import PersonaForm
from .models import Estudiante

def formulario_view(request):
    if request.method == 'POST':
        form = PersonaForm(request.POST)
        if form.is_valid():
            cleaned_data = form.cleaned_data
            
            # Default to False if 'trabajas' is not in cleaned_data
            trabaja = cleaned_data.get('trabajas', False)
            
            # Create an instance of Estudiante with the cleaned data
            estudiante = Estudiante(
                nombre=cleaned_data['nombre'],
                edad=cleaned_data['edad'],
                sexo=cleaned_data['sexo'],
                fecha_nacimiento=cleaned_data['fecha_nacimiento'],
                lugar_nacimiento=cleaned_data['lugar_nacimiento'],
                numero_matricula=cleaned_data['numero_matricula'],
                direccion=cleaned_data['direccion'],
                telefono=cleaned_data['telefono'],
                celular=cleaned_data['celular'],
                telefono_emergencia=cleaned_data['telefono_emergencia'],
                carrera=cleaned_data['carrera'],
                cuatrimestre=cleaned_data['cuatrimestre'],
                grupo_escolar=cleaned_data['grupo_escolar'],
                estado_civil=cleaned_data['estado_civil'],
                estado_civil_otro=cleaned_data.get('estado_civil_otro'),
                nombre_tutor=cleaned_data['nombre_tutor'],
                numero_seguro_social=cleaned_data['numero_seguro_social'],
                correo_electronico=cleaned_data['correo_electronico'],
                tipo_sangre=cleaned_data['tipo_sangre'],
                tipo_religion=cleaned_data['tipo_religion'],
                lugar_procedencia=cleaned_data['lugar_procedencia'],
                con_quien_vives=cleaned_data['con_quien_vives'],
                otros_familiares=cleaned_data.get('otros_familiares'),
                trabajas=trabaja,
                tipo_empresa=cleaned_data.get('tipo_empresa'),
                horas_trabajo=cleaned_data.get('horas_trabajo'),
                ingreso_mensual=cleaned_data.get('ingreso_mensual'),
                horario_laboral=cleaned_data.get('horario_laboral'),
                dependes_economicamente=cleaned_data['dependes_economicamente'],
                otros_dependientes=cleaned_data.get('otros_dependientes'),
                ocupacion_padre=cleaned_data['ocupacion_padre'],
                ingreso_padre=cleaned_data.get('ingreso_padre'),
                ocupacion_madre=cleaned_data['ocupacion_madre'],
                ingreso_madre=cleaned_data.get('ingreso_madre'),
                hermanos=cleaned_data.get('hermanos'),
                lugar_entre_hermanos=cleaned_data.get('lugar_entre_hermanos'),
                actividad_hermanos=cleaned_data.get('actividad_hermanos'),
                tipo_casa=cleaned_data['tipo_casa'],
                tipo_casa_otros=cleaned_data.get('tipo_casa_otros'),
                beca_bachillerato=cleaned_data['beca_bachillerato'],
                tipo_beca_bachillerato=cleaned_data.get('tipo_beca_bachillerato'),
                requieres_beca=cleaned_data['requieres_beca'],
                motivo_beca=cleaned_data.get('motivo_beca'),
                cualidades=cleaned_data.get('cualidades'),
                habilidades=cleaned_data.get('habilidades'),
                debilidades=cleaned_data.get('debilidades'),
                valores=cleaned_data.get('valores'),
                disgustos=cleaned_data.get('disgustos'),
                temores=cleaned_data.get('temores'),
                pareja_sentimental=cleaned_data.get('pareja_sentimental'),
                planes_matrimonio=cleaned_data.get('planes_matrimonio'),
                futuro_personal=cleaned_data.get('futuro_personal'),
                futuro_academico=cleaned_data.get('futuro_academico'),
                futuro_profesional=cleaned_data.get('futuro_profesional'),
                problemas_personales=cleaned_data.get('problemas_personales'),
                tiempo_libre=cleaned_data.get('tiempo_libre'),
                motivacion_estudio=cleaned_data['motivacion_estudio'],
                motivo_motivacion=cleaned_data.get('motivo_motivacion'),
                transporte_universidad=cleaned_data.get('transporte_universidad'),
                estado_salud=cleaned_data.get('estado_salud'),
                tratamiento_medico=cleaned_data['tratamiento_medico'],
                tipo_tratamiento=cleaned_data.get('tipo_tratamiento'),
                discapacidad=cleaned_data['discapacidad'],
                tipo_discapacidad=cleaned_data.get('tipo_discapacidad'),
                ayuda_psicologica=cleaned_data['ayuda_psicologica'],
                motivo_ayuda=cleaned_data.get('motivo_ayuda'),
                dependientes_economicos=cleaned_data.get('dependientes_economicos'),
                escuela_procedencia=cleaned_data['escuela_procedencia'],
                especialidad_promedio=cleaned_data['especialidad_promedio'],
                motivo_universidad=cleaned_data.get('motivo_universidad'),
                universidad_primera_opcion=cleaned_data['universidad_primera_opcion'],
                motivo_universidad_opcion=cleaned_data.get('motivo_universidad_opcion'),
                carrera_primera_opcion=cleaned_data['carrera_primera_opcion'],
                motivo_carrera_opcion=cleaned_data.get('motivo_carrera_opcion'),
                expectativas_carrera=cleaned_data.get('expectativas_carrera'),
                examen_admision=cleaned_data['examen_admision'],
                cual_otra_escuela=cleaned_data.get('cual_otra_escuela'),
                materias_dificultad=cleaned_data.get('materias_dificultad'),
                materias_reprobadas=cleaned_data['materias_reprobadas'],
                cuales_materias_reprobadas=cleaned_data.get('cuales_materias_reprobadas'),
                materias_facil=cleaned_data.get('materias_facil'),
                forma_trabajar=cleaned_data['forma_trabajar'],
                motivo_forma_trabajar=cleaned_data.get('motivo_forma_trabajar'),
                tecnica_estudio=cleaned_data['tecnica_estudio'],
                cual_tecnica_estudio=cleaned_data.get('cual_tecnica_estudio'),
                espacio_estudio=cleaned_data['espacio_estudio'],
                cual_espacio_estudio=cleaned_data.get('cual_espacio_estudio'),
                libros_apoyo=cleaned_data['libros_apoyo'],
                cantidad_libros=cleaned_data.get('cantidad_libros'),
                computadora_internet=cleaned_data['computadora_internet'],
                tipo_computadora=cleaned_data.get('tipo_computadora'),
                tipo_internet=cleaned_data.get('tipo_internet'),
                planes_terminar_carrera=cleaned_data.get('planes_terminar_carrera'),
                tipo_trabajo=cleaned_data.get('tipo_trabajo')
            )

            # Save the Estudiante instance to the database
            estudiante.save()

            # You can pass an empty form after sending the data
            return render(request, 'formulario.html', {'form': PersonaForm(), 'enviado': True})
    else:
        form = PersonaForm()

    return render(request, 'formulario.html', {'form': form})
 


from django.shortcuts import render
from django.http import HttpResponse
from .models import Grupo, Profesor, Estudiante
import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

def lista_grupos(request):
    profesores = Profesor.objects.all()
    grupos = Grupo.objects.all()
    estudiantes = Estudiante.objects.none()

    if 'profesor_id' in request.GET:
        profesor_id = request.GET['profesor_id']
        if profesor_id:
            profesor = Profesor.objects.get(id=profesor_id)
            grupos = Grupo.objects.filter(tutor=profesor)
            estudiantes = Estudiante.objects.filter(grupo_escolar__in=grupos, carrera=profesor.carrera)

    if 'export' in request.GET and estudiantes.exists():
        return exportar_estudiantes_a_excel(estudiantes)

    return render(request, 'lista_grupos.html', {'grupos': grupos, 'profesores': profesores, 'estudiantes': estudiantes})

def exportar_estudiantes_a_excel(estudiantes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    columnas = ['Nombre Completo ', 'Edad', 'Número de Matrícula', 'Carrera', 'Grupo Escolar']
    for col_num, column_title in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = column_title

    for row_num, estudiante in enumerate(estudiantes, 2):
        ws[f'A{row_num}'] = estudiante.nombre
        ws[f'B{row_num}'] = estudiante.edad
        ws[f'C{row_num}'] = estudiante.numero_matricula
        ws[f'D{row_num}'] = estudiante.carrera.nombre
        ws[f'E{row_num}'] = estudiante.grupo_escolar.nombre

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estudiantes.xlsx'
    wb.save(response)
    return response


# mi_aplicacion/views.py

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Profesor, Carrera, Grupo
from .forms import ProfesorForm

def pantalla_administracion(request):
    profesor_id = request.GET.get('profesor_id')
    if request.method == 'POST':
        if profesor_id:
            profesor = get_object_or_404(Profesor, pk=profesor_id)
            form = ProfesorForm(request.POST, instance=profesor)
        else:
            form = ProfesorForm(request.POST)
        
        if form.is_valid():
            form.save()
            # Redirigir o actualizar la página según sea necesario
    else:
        if profesor_id:
            profesor = get_object_or_404(Profesor, pk=profesor_id)
            form = ProfesorForm(instance=profesor)
        else:
            form = ProfesorForm()
    
    profesores = Profesor.objects.all()
    return render(request, 'pantalla_administracion.html', {'form': form, 'profesores': profesores})

def get_grupos(request, carrera_id):
    grupos = Grupo.objects.filter(carrera_id=carrera_id).order_by('nombre').values('id', 'nombre')
    return JsonResponse(list(grupos), safe=False)

from django.shortcuts import render, get_object_or_404, redirect
from .models import Carrera, Grupo
from .forms import CarreraForm, GrupoForm

def administrar_carreras_grupos(request):
    carreras = Carrera.objects.all()
    grupos = Grupo.objects.all()

    carrera_form = CarreraForm(request.POST or None)
    grupo_form = GrupoForm(request.POST or None)

    if request.method == 'POST':
        if 'carrera_submit' in request.POST and carrera_form.is_valid():
            carrera_form.save()
            return redirect('administrar_carreras_grupos')

        if 'grupo_submit' in request.POST and grupo_form.is_valid():
            grupo_form.save()
            return redirect('administrar_carreras_grupos')

    if 'carrera_id' in request.GET:
        carrera_id = request.GET.get('carrera_id')
        carrera = get_object_or_404(Carrera, id=carrera_id)
        carrera_form = CarreraForm(instance=carrera)

    if 'grupo_id' in request.GET:
        grupo_id = request.GET.get('grupo_id')
        grupo = get_object_or_404(Grupo, id=grupo_id)
        grupo_form = GrupoForm(instance=grupo)

    return render(request, 'administrar_carreras_grupos.html', {
        'carrera_form': carrera_form,
        'grupo_form': grupo_form,
        'carreras': carreras,
        'grupos': grupos,
    })










